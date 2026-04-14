"""Dev-only helper: extract seed reference data from the Diretrizes XLSX into
Python literals suitable for inclusion in an Alembic data migration.

Usage:
    uv run python scripts/extract_reference_seed.py > /tmp/reference_seed_literals.py

The XLSX lives under docs/Documentos de Instrução/ and is not shipped with
the deployed backend. This script is only used to regenerate the literals
embedded in the data migration when the upstream spreadsheet changes.
"""

from __future__ import annotations

import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = (
    Path(__file__).resolve().parents[2]
    / "docs"
    / "Documentos de Instrução"
    / "03) Diretrizes para a configuração de IA - Relatórios de Sustentabilidade.xlsx"
)


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = "".join(char for char in normalized if not unicodedata.combining(char))
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_only).strip("-").lower()
    return slug


_SESSAO_TO_DIRECTORY = {
    "visao e estrategia": "visao-estrategica-de-sustentabilidade",
    "governanca corporativa": "governanca-corporativa",
    "gestao ambiental": "gestao-ambiental",
    "desempenho social": "desempenho-social",
    "gestao e desempenho economico": "gestao-de-desempenho-economico",
    "gestao economica": "gestao-de-desempenho-economico",
    "desempenho economico": "gestao-de-desempenho-economico",
    "relacionamento com stakeholders": "relacionamento-com-stakeholders",
    "inovacao e desenvolvimento": "inovacao-e-desenvolvimento-tecnologico",
    "inovacao": "inovacao-e-desenvolvimento-tecnologico",
    "relatorios e normas": "relatorios-e-normas",
    "comunicacao e transparencia": "comunicacao-e-transparencia",
    "auditorias e avaliacoes": "auditorias-e-avaliacoes",
}


def _normalize_sessao(value: str) -> str:
    slug = _slugify(value).replace("-", " ")
    for key, directory in _SESSAO_TO_DIRECTORY.items():
        if slug.startswith(key):
            return directory
    return slug


def _extract_gri_code(raw: str) -> tuple[str, str] | None:
    """Return (code, family) or None if the cell is a section header."""
    if not raw:
        return None
    normalized = raw.strip()
    match = re.match(r"^GRI\s+(\d+)-(\d+[a-z]?)$", normalized, re.IGNORECASE)
    if not match:
        return None
    family_num = int(match.group(1))
    if family_num < 10:
        family = str(family_num)
    else:
        family = str((family_num // 100) * 100)
    return f"GRI {family_num}-{match.group(2)}", family


def _load_workbook():
    if not XLSX_PATH.exists():
        raise SystemExit(f"XLSX not found: {XLSX_PATH}")
    return load_workbook(XLSX_PATH, data_only=True)


def dump_gri_standards(wb) -> list[dict]:
    ws = wb["Índice GRI"]
    out = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        code_cell = row[0]
        text_cell = row[1]
        if not code_cell:
            continue
        extracted = _extract_gri_code(str(code_cell))
        if not extracted:
            continue
        code, family = extracted
        if code in seen:
            continue
        seen.add(code)
        standard_text = str(text_cell).strip() if text_cell else ""
        # strip trailing page numbers like "... 7" that appear in some rows
        standard_text = re.sub(r"\s+\d+\s*$", "", standard_text)
        out.append({"code": code, "family": family, "standard_text": standard_text})
    return out


def dump_ods(wb) -> tuple[list[dict], list[dict]]:
    ws = wb["ODS Pacto Global e Agenda 2030"]
    goals: dict[int, str] = {}
    metas: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ods_cell, objetivo_cell, meta_cell = row[0], row[1], row[2]
        if not ods_cell:
            continue
        match = re.match(r"^ODS\s+(\d+)$", str(ods_cell).strip())
        if not match:
            continue
        ods_number = int(match.group(1))
        objetivo = str(objetivo_cell).strip() if objetivo_cell else ""
        if ods_number not in goals:
            goals[ods_number] = objetivo
        if meta_cell:
            meta_text = str(meta_cell).strip()
            meta_match = re.match(r"^(\d+\.[0-9a-z]+)\s+(.*)$", meta_text)
            if meta_match:
                metas.append(
                    {
                        "ods_number": ods_number,
                        "meta_code": meta_match.group(1),
                        "meta_text": meta_match.group(2).rstrip("."),
                    }
                )
            else:
                metas.append(
                    {
                        "ods_number": ods_number,
                        "meta_code": "",
                        "meta_text": meta_text,
                    }
                )
    goals_out = [{"ods_number": n, "objetivo": goals[n]} for n in sorted(goals)]
    return goals_out, metas


def dump_captacao(wb) -> list[dict]:
    ws = wb["Matriz de captação de dados"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sessao, tipo_dado, gri_code, descricao, fonte, evidencia = (
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
        )
        if not sessao or not tipo_dado:
            continue
        normalized_gri = None
        if gri_code:
            extracted = _extract_gri_code(str(gri_code))
            if extracted:
                normalized_gri = extracted[0]
        out.append(
            {
                "sessao": _normalize_sessao(str(sessao)),
                "tipo_dado": str(tipo_dado).strip(),
                "gri_code": normalized_gri,
                "descricao": str(descricao).strip() if descricao else "",
                "fonte_documental": str(fonte).strip() if fonte else "",
                "tipo_evidencia": str(evidencia).strip() if evidencia else "",
            }
        )
    return out


def dump_indicators(wb) -> list[dict]:
    ws = wb["Indicadores ESG"]
    out = []
    current_tema = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        tema_cell, indicador_cell, unidade_cell = row[0], row[1], row[2]
        if tema_cell:
            current_tema = str(tema_cell).strip()
        if not indicador_cell or not current_tema:
            continue
        out.append(
            {
                "tema": current_tema,
                "indicador": str(indicador_cell).strip(),
                "unidade": str(unidade_cell).strip() if unidade_cell else "",
            }
        )
    return out


def _format_literal(name: str, rows: list[dict]) -> str:
    lines = [f"{name} = ["]
    for row in rows:
        items = ", ".join(f"{k!r}: {v!r}" for k, v in row.items())
        lines.append("    {" + items + "},")
    lines.append("]")
    return "\n".join(lines)


def main() -> int:
    wb = _load_workbook()
    gri = dump_gri_standards(wb)
    goals, metas = dump_ods(wb)
    captacao = dump_captacao(wb)
    indicators = dump_indicators(wb)

    sys.stdout.write(f"# Generated from {XLSX_PATH.name}\n")
    sys.stdout.write(f"# GRI codes: {len(gri)}\n")
    sys.stdout.write(f"# ODS goals: {len(goals)}\n")
    sys.stdout.write(f"# ODS metas: {len(metas)}\n")
    sys.stdout.write(f"# Captacao rows: {len(captacao)}\n")
    sys.stdout.write(f"# Indicator templates: {len(indicators)}\n\n")
    sys.stdout.write(_format_literal("GRI_STANDARDS", gri) + "\n\n")
    sys.stdout.write(_format_literal("ODS_GOALS", goals) + "\n\n")
    sys.stdout.write(_format_literal("ODS_METAS", metas) + "\n\n")
    sys.stdout.write(_format_literal("CAPTACAO_ROWS", captacao) + "\n\n")
    sys.stdout.write(_format_literal("INDICATOR_TEMPLATES", indicators) + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
