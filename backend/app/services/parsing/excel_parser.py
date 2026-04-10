from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook

from app.services.parsing import ParsedDocumentResult


def _normalize_cell(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None

    return str(value)


def _detect_header(
    rows: list[list[str | None]],
) -> tuple[list[str | None], list[list[str | None]]]:
    first_non_empty = next(
        (row for row in rows if any(cell not in {None, ""} for cell in row)),
        [],
    )
    if not first_non_empty:
        return [], []

    start_index = rows.index(first_non_empty)
    return first_non_empty, rows[start_index + 1 :]


def parse_xlsx_document(file_bytes: bytes) -> ParsedDocumentResult:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheets_payload: list[dict[str, Any]] = []
    extracted_sections: list[str] = []

    for worksheet in workbook.worksheets:
        normalized_rows = [
            [_normalize_cell(value) for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        filtered_rows = [
            row
            for row in normalized_rows
            if any(cell not in {None, ""} for cell in row)
        ]
        header, rows = _detect_header(filtered_rows)
        sheets_payload.append(
            {
                "sheet_name": worksheet.title,
                "header": header,
                "rows": rows,
            }
        )

        extracted_sections.append(f"Aba: {worksheet.title}")
        if header:
            extracted_sections.append(" | ".join(cell or "" for cell in header).strip())
        for row in rows:
            extracted_sections.append(" | ".join(cell or "" for cell in row).strip())

    return ParsedDocumentResult(
        extracted_text="\n".join(
            section for section in extracted_sections if section
        ).strip(),
        parsed_payload={"format": "xlsx", "sheets": sheets_payload},
    )


def parse_csv_document(file_bytes: bytes) -> ParsedDocumentResult:
    decoded = file_bytes.decode("utf-8-sig")
    sample = decoded[:1024]
    dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
    reader = csv.reader(StringIO(decoded), dialect=dialect)
    normalized_rows = [[_normalize_cell(cell) for cell in row] for row in reader]
    filtered_rows = [
        row for row in normalized_rows if any(cell not in {None, ""} for cell in row)
    ]
    header, rows = _detect_header(filtered_rows)

    extracted_lines = []
    if header:
        extracted_lines.append(" | ".join(cell or "" for cell in header).strip())
    for row in rows:
        extracted_lines.append(" | ".join(cell or "" for cell in row).strip())

    return ParsedDocumentResult(
        extracted_text="\n".join(line for line in extracted_lines if line).strip(),
        parsed_payload={
            "format": "csv",
            "header": header,
            "rows": rows,
        },
    )
